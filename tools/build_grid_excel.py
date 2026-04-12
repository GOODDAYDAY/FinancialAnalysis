"""
build_grid_excel.py — Grid Trading Excel Calculator v2

Layout (8 rows header + 20 data rows):
  Row 1  : Title + stock name input
  Rows 2-6: Parameters (2-column, 5 rows × 2 = 10 params)
  Row 7  : Summary stats (single row, multi-column)
  Row 8  : Grid column headers
  Rows 9-28: 20 grid levels (auto-expand based on upper/lower limit)

Grid columns (A-L, 12 total):
  A 序号  B 网格价格  C 操作  D 成交金额  E 手续费  F 净利/格
  G 累计持仓  H 持仓均价  I 持仓市值  J 资金余额  K 总资产  L 总盈亏

Run:  python tools/build_grid_excel.py
Out:  tools/grid_trading_calculator.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

# ─── Layout ──────────────────────────────────────────────────────────────────
TITLE_ROW    = 1
PARAM_START  = 2    # parameter rows 2-6
PARAM_END    = 6
SUMMARY_ROW  = 7
GRID_HDR_ROW = 8
DATA_START   = 9
DATA_END     = 28   # 20 grid rows

# Parameter cell addresses  (left column = E, right column = K)
P_CUR  = "$E$2"   # 当前价格
P_INT  = "$E$3"   # 网格间距
P_SHR  = "$E$4"   # 每格交易股数
P_FEE  = "$E$5"   # 固定手续费/笔
P_DN   = "$E$6"   # 网格下限
P_RATE = "$K$2"   # 比例费率
P_ISHP = "$K$3"   # 初始持仓(股)
P_UP   = "$K$4"   # 网格上限
P_CAP  = "$K$5"   # 初始资金
# $K$6 = stock name (display only, used in title)

# Grid column indices (1-based)
C_NO   = 1    # A  序号
C_PRC  = 2    # B  网格价格
C_ACT  = 3    # C  操作
C_AMT  = 4    # D  成交金额
C_COM  = 5    # E  手续费(单腿)
C_NET  = 6    # F  净利/格
C_CSHP = 7    # G  累计持仓
C_AVG  = 8    # H  持仓均价
C_MKT  = 9    # I  持仓市值
C_CASH = 10   # J  资金余额
C_TOT  = 11   # K  总资产
C_PNL  = 12   # L  总盈亏

# ─── Styles ──────────────────────────────────────────────────────────────────

def _fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def _font(bold=False, size=10, color="000000", italic=False, white=False):
    return Font(bold=bold, size=size, color="FFFFFF" if white else color, italic=italic)
def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

BORDER  = _border()
ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_L = Alignment(horizontal="left",   vertical="center")
ALIGN_R = Alignment(horizontal="right",  vertical="center")
ALIGN_W = Alignment(horizontal="center", vertical="center", wrap_text=True)

F_TITLE  = _fill("1F4E79")   # dark blue
F_PLAB   = _fill("D6E4F0")   # light blue — label
F_PVAL   = _fill("FFFACD")   # yellow — editable input
F_SUM    = _fill("EEF2F7")   # very light blue — summary
F_GHDR   = _fill("2E75B6")   # blue — grid header
F_YELLOW = _fill("FFFF00")   # CURRENT row
F_ORANGE = _fill("FFC000")   # net profit ≤ 0 warning
F_LGRN   = _fill("C6EFCE")   # BUY action
F_LRED   = _fill("FFC7CE")   # SELL action
F_DGRN   = _fill("375623")   # dark green fill for positive P&L font
F_DRED   = _fill("9C0006")   # dark red fill for negative P&L font

# ─── Main ────────────────────────────────────────────────────────────────────

def build(output_path: str) -> None:
    """Generate the grid trading calculator Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "网格计算"

    _set_col_widths(ws)
    _write_title(ws)
    _write_params(ws)
    _write_summary(ws)
    _write_grid_header(ws)
    _write_grid_data(ws)
    _apply_cond_fmt(ws)
    ws.freeze_panes = f"A{DATA_START}"

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"[OK] {output_path}")

# ─── Section writers ─────────────────────────────────────────────────────────

def _set_col_widths(ws) -> None:
    for col, w in zip("ABCDEFGHIJKL",
                      [7, 12, 10, 13, 12, 13, 13, 13, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w


def _write_title(ws) -> None:
    """Row 1: Title with embedded stock name reference."""
    ws.row_dimensions[TITLE_ROW].height = 34

    # Merged title  A1:I1  — formula embeds stock name from K6
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = '="网格交易计算器  ·  "&$K$6'
    c.font      = _font(bold=True, size=15, white=True)
    c.fill      = F_TITLE
    c.alignment = ALIGN_L

    # Stock name label  J1
    j1 = ws["J1"]
    j1.value     = "标的名称"
    j1.font      = _font(bold=True, white=True)
    j1.fill      = F_TITLE
    j1.alignment = ALIGN_R

    # Stock name input  K1:L1 (yellow)
    ws.merge_cells("K1:L1")
    k1 = ws["K1"]
    k1.value     = "泰达股份"
    k1.font      = _font(bold=True, size=11)
    k1.fill      = F_PVAL
    k1.alignment = ALIGN_C
    k1.border    = BORDER

    # Title note — reference K6 is actually K1 now (reassign P_NAME to K1)
    # We keep K6 free for the 5th right-param row used as stock name alias
    # Actually: stock name is in K1, we'll reference it via $K$1 in the title formula above


def _write_params(ws) -> None:
    """Rows 2-6: two-column parameter layout.

    Left (label A:D merged, value E):  current price, interval, shares, fixed fee, lower limit
    Right (label G:J merged, value K): fee rate, initial shares, upper limit, initial capital, (blank)
    Column F, L: spacers
    """
    left_params = [
        ("当前价格  Current Price  (元)",   4.10,   "#,##0.000"),
        ("网格间距  Grid Interval  (元)",   0.20,   "#,##0.000"),
        ("每格股数  Shares per Grid  (股)", 1000,   "#,##0"),
        ("固定手续费  Fixed Fee  (元/笔)",  10.0,   "#,##0.00"),
        ("网格下限  Lower Limit  (元)",     3.50,   "#,##0.000"),
    ]
    right_params = [
        ("比例费率  Fee Rate  (如万分之一=0.0001)", 0.0001, "0.0000%"),
        ("初始持仓  Initial Shares  (股)",          5000,   "#,##0"),
        ("网格上限  Upper Limit  (元)",              5.00,   "#,##0.000"),
        ("初始资金  Initial Capital  (元)",          50000,  "#,##0"),
        ("",  "",  None),   # row 6 right side left empty
    ]

    for i, ((llabel, lval, lfmt), (rlabel, rval, rfmt)) in \
            enumerate(zip(left_params, right_params)):
        r = PARAM_START + i
        ws.row_dimensions[r].height = 21

        # Left label (A:D merged)
        ws.merge_cells(f"A{r}:D{r}")
        lc = ws[f"A{r}"]
        lc.value, lc.fill, lc.font, lc.alignment, lc.border = \
            llabel, F_PLAB, _font(bold=True, size=9), ALIGN_L, BORDER

        # Left value (E)
        lv = ws.cell(r, 5, lval)
        lv.fill, lv.font, lv.alignment, lv.border = F_PVAL, _font(bold=True), ALIGN_R, BORDER
        if lfmt: lv.number_format = lfmt

        # Right label (G:J merged)
        if rlabel:
            ws.merge_cells(f"G{r}:J{r}")
            rc = ws[f"G{r}"]
            rc.value, rc.fill, rc.font, rc.alignment, rc.border = \
                rlabel, F_PLAB, _font(bold=True, size=9), ALIGN_L, BORDER
            rv = ws.cell(r, 11, rval)
            rv.fill, rv.font, rv.alignment, rv.border = F_PVAL, _font(bold=True), ALIGN_R, BORDER
            if rfmt: rv.number_format = rfmt

    # Tip text in the spacer area
    ws.cell(PARAM_END, 7).value = \
        "← 修改黄色格，表格自动更新  |  Edit yellow cells to recalculate"
    ws.cell(PARAM_END, 7).font = _font(italic=True, size=8, color="888888")
    ws.merge_cells(f"G{PARAM_END}:L{PARAM_END}")


def _write_summary(ws) -> None:
    """Row 7: Compact single-row summary statistics."""
    ws.row_dimensions[SUMMARY_ROW].height = 22
    r = SUMMARY_ROW
    DS, DE = DATA_START, DATA_END

    def stat(col_label, col_val, label, formula, fmt=None, merge_val=None):
        """Write a (label, value) pair in the summary row."""
        lc = ws.cell(r, col_label, label)
        lc.fill, lc.font, lc.alignment, lc.border = \
            F_SUM, _font(bold=True, size=9, color="1F4E79"), ALIGN_R, BORDER

        vc = ws.cell(r, col_val, formula)
        vc.fill, vc.font, vc.alignment, vc.border = \
            F_SUM, _font(size=9), ALIGN_L, BORDER
        if fmt: vc.number_format = fmt

        if merge_val:
            ws.merge_cells(merge_val)

    # A7: total grids
    stat(1, 2, "总格数",
         f'=SUMPRODUCT((B{DS}:B{DE}<>"")*1)',
         "#,##0")

    # C7: net profit per grid
    stat(3, 4, "净利/格",
         f'={P_INT}*{P_SHR}-2*MAX({P_FEE},{P_CUR}*{P_SHR}*{P_RATE})',
         "#,##0.00")

    # E7: min required capital (label) + F7:G7 merged (value)
    lc = ws.cell(r, 5, "最低资金")
    lc.fill, lc.font, lc.alignment, lc.border = \
        F_SUM, _font(bold=True, size=9, color="1F4E79"), ALIGN_R, BORDER

    ws.merge_cells(f"F{r}:G{r}")
    vc = ws.cell(r, 6,
         f'={P_DN}*{P_SHR}*COUNTIF(C{DS}:C{DE},"BUY")')
    vc.fill, vc.font, vc.alignment, vc.border = \
        F_SUM, _font(size=9), ALIGN_L, BORDER
    vc.number_format = "#,##0"

    # H7: capital sufficiency (label) + I7:J7 merged (value)
    lc2 = ws.cell(r, 8, "资金充足")
    lc2.fill, lc2.font, lc2.alignment, lc2.border = \
        F_SUM, _font(bold=True, size=9, color="1F4E79"), ALIGN_R, BORDER

    ws.merge_cells(f"I{r}:J{r}")
    vc2 = ws.cell(r, 9,
          f'=IF({P_CAP}>=F{r},'
          f'"✓ 充裕  余额 "&TEXT({P_CAP}-F{r},"#,##0")&" 元",'
          f'"✗ 不足  缺口 "&TEXT(F{r}-{P_CAP},"#,##0")&" 元")')
    vc2.fill, vc2.font, vc2.alignment, vc2.border = \
        F_SUM, _font(size=9), ALIGN_L, BORDER

    # K7: break-even label + L7 value
    lc3 = ws.cell(r, 11, "盈亏平衡")
    lc3.fill, lc3.font, lc3.alignment, lc3.border = \
        F_SUM, _font(bold=True, size=9, color="1F4E79"), ALIGN_R, BORDER

    vc3 = ws.cell(r, 12,
          f'=IF(D{r}<=0,"净利为负",'
          f'TEXT(CEILING(SUMIF(B{DS}:B{DE},"<>",E{DS}:E{DE})*2/D{r},1),"0")&"次振荡")')
    vc3.fill, vc3.font, vc3.alignment, vc3.border = \
        F_SUM, _font(size=9), ALIGN_L, BORDER


def _write_grid_header(ws) -> None:
    """Row 8: Grid table column headers."""
    ws.row_dimensions[GRID_HDR_ROW].height = 44
    headers = [
        "序号\nNo.",
        "网格价格\nPrice (元)",
        "操作\nAction",
        "成交金额\nAmount (元)",
        "手续费\nComm.(元)",
        "净利/格\nNet P&L(元)",
        "累计持仓\nCum.Shares",
        "持仓均价\nAvg Cost(元)",
        "持仓市值\nMkt Val(元)",
        "资金余额\nCash(元)",
        "总资产\nTotal Val(元)",
        "总盈亏\nP&L(元)",
    ]
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(GRID_HDR_ROW, col, hdr)
        c.fill, c.font, c.alignment, c.border = \
            F_GHDR, _font(bold=True, white=True), ALIGN_W, BORDER


def _write_grid_data(ws) -> None:
    """Rows 9-28: One formula row per pre-built grid level."""
    DS = DATA_START   # 9

    # k_expr: how many grid levels above DS this row is relative to the CURRENT level.
    #   k > 0  → below current price (BUY grids, price goes down)
    #   k = 0  → CURRENT level (price equals current price)
    #   k < 0  → above current price (SELL grids, price goes up)
    k = f"(ROUND(({P_CUR}-{P_DN})/{P_INT},0)-(ROW()-{DS}))"

    for r in range(DS, DATA_END + 1):
        b  = f"B{r}"    # grid price
        e  = f"E{r}"    # commission per leg
        g  = f"G{r}"    # cumulative shares
        i_ = f"I{r}"    # market value
        f_ = f"F{r}"    # net profit per grid
        j  = f"J{r}"    # cash remaining

        blank = f'=IF({b}="","",'  # prefix: output "" for empty (out-of-range) rows

        # A: Grid sequence number
        ws.cell(r, C_NO).value = f'=IF({b}="","",ROW()-{DS - 1})'

        # B: Grid Price = lower_limit + row_index × interval, capped at upper_limit
        ws.cell(r, C_PRC).value = (
            f'=IF(ROUND({P_DN}+(ROW()-{DS})*{P_INT},4)<={P_UP},'
            f'ROUND({P_DN}+(ROW()-{DS})*{P_INT},4),"")'
        )
        ws.cell(r, C_PRC).number_format = "#,##0.000"

        # C: Action  — compare with ±tolerance to handle floating point
        ws.cell(r, C_ACT).value = (
            f'=IF({b}="","",IF(ROUND({b},4)<ROUND({P_CUR},4),"BUY",'
            f'IF(ROUND({b},4)>ROUND({P_CUR},4),"SELL","CURRENT")))'
        )

        # D: Trade Amount = price × shares_per_grid
        ws.cell(r, C_AMT).value = f'{blank}{b}*{P_SHR})'
        ws.cell(r, C_AMT).number_format = "#,##0.00"

        # E: Commission per leg = max(fixed_fee, amount × fee_rate)
        ws.cell(r, C_COM).value = f'{blank}MAX({P_FEE},D{r}*{P_RATE}))'
        ws.cell(r, C_COM).number_format = "#,##0.00"

        # F: Net Profit per round-trip = interval × shares − 2 × commission
        # CURRENT row has no completed round-trip → N/A
        ws.cell(r, C_NET).value = (
            f'{blank}IF(C{r}="CURRENT","N/A",{P_INT}*{P_SHR}-2*MAX({P_FEE},D{r}*{P_RATE})))'
        )
        ws.cell(r, C_NET).number_format = "#,##0.00"

        # G: Cumulative Shares = initial_shares + shares_per_grid × k
        #   At CURRENT (k=0): = initial_shares
        #   BUY grids (k>0):  initial + k × shares (more shares held)
        #   SELL grids (k<0): initial + k × shares (fewer shares held, k negative)
        ws.cell(r, C_CSHP).value = f'{blank}{P_ISHP}+{P_SHR}*{k})'
        ws.cell(r, C_CSHP).number_format = "#,##0"

        # H: Avg Cost Price (weighted average of all purchases)
        #   BUY/CURRENT (k >= 0):
        #     cost = initial_shares × current_price
        #           + shares × k × (current_price − interval × (k+1)/2)
        #     shares_total = initial_shares + shares × k
        #     avg = cost / shares_total
        #   SELL (k < 0): avg cost of remaining shares is unchanged → current_price
        ws.cell(r, C_AVG).value = (
            f'{blank}IF({k}>=0,'
            f'({P_ISHP}*{P_CUR}+{P_SHR}*{k}*({P_CUR}-{P_INT}*({k}+1)/2))/({P_ISHP}+{P_SHR}*{k}),'
            f'{P_CUR}))'
        )
        ws.cell(r, C_AVG).number_format = "#,##0.000"

        # I: Market Value = cumulative_shares × grid_price
        ws.cell(r, C_MKT).value = f'{blank}{g}*{b})'
        ws.cell(r, C_MKT).number_format = "#,##0.00"

        # J: Cash Remaining — analytical formula
        #
        # For k BUY trades (k > 0), buying at prices:
        #   p1 = current−1×interval, p2 = current−2×interval, …, pk = current−k×interval
        #   avg_buy_price = current − interval × (k+1)/2
        #   total_spent = shares × k × avg_buy_price + k × commission_approx
        #   cash = initial_capital − total_spent
        #
        # For k_sell SELL trades (k < 0, abs_k = −k), selling at prices:
        #   q1 = current+1×interval, …, qk_sell = current+k_sell×interval
        #   avg_sell_price = current + interval × (abs_k+1)/2
        #   total_received = shares × abs_k × avg_sell_price − abs_k × commission_approx
        #   cash = initial_capital + total_received
        #
        # Commission approximated using average price for each direction.
        # (For typical fixed_fee=10 >> price×shares×rate≈0.4, error is negligible.)

        avg_buy  = f"({P_CUR}-{P_INT}*({k}+1)/2)"
        avg_sell = f"({P_CUR}+{P_INT}*(-{k}+1)/2)"

        ws.cell(r, C_CASH).value = (
            f'{blank}IF({k}>=0,'
            # BUY direction
            f'{P_CAP}-{P_SHR}*{k}*{avg_buy}-{k}*MAX({P_FEE},{avg_buy}*{P_SHR}*{P_RATE}),'
            # SELL direction
            f'{P_CAP}+{P_SHR}*(-{k})*{avg_sell}-(-{k})*MAX({P_FEE},{avg_sell}*{P_SHR}*{P_RATE})'
            f'))'
        )
        ws.cell(r, C_CASH).number_format = "#,##0.00"

        # K: Total Portfolio Value = Cash + Market Value
        ws.cell(r, C_TOT).value = f'{blank}{j}+{i_})'
        ws.cell(r, C_TOT).number_format = "#,##0.00"

        # L: Total P&L vs Initial State
        #   Initial value = initial_capital (cash) + initial_shares × current_price (shares)
        #   P&L = Total Value − Initial Value
        ws.cell(r, C_PNL).value = (
            f'{blank}K{r}-({P_ISHP}*{P_CUR}+{P_CAP}))'
        )
        ws.cell(r, C_PNL).number_format = '+#,##0.00;-#,##0.00;0.00'

        # ── Row styling ──
        for col in range(1, 13):
            c = ws.cell(r, col)
            c.border    = BORDER
            c.font      = _font(size=9)
            c.alignment = ALIGN_C if col in (C_NO, C_ACT) else ALIGN_R
        ws.row_dimensions[r].height = 18


def _apply_cond_fmt(ws) -> None:
    """Conditional formatting rules for the grid data range."""
    DS, DE = DATA_START, DATA_END
    all_  = f"A{DS}:L{DE}"
    col_c = f"C{DS}:C{DE}"
    col_g = f"G{DS}:G{DE}"
    col_l = f"L{DS}:L{DE}"

    # CURRENT row → yellow (full row)
    ws.conditional_formatting.add(all_, FormulaRule(
        formula=[f'$C{DS}="CURRENT"'], fill=F_YELLOW))

    # Net profit per grid ≤ 0 → orange warning (rows where F is a negative number)
    ws.conditional_formatting.add(all_, FormulaRule(
        formula=[f'AND(ISNUMBER($F{DS}),$F{DS}<=0)'], fill=F_ORANGE))

    # Cumulative shares < 0 → red fill on column G (over-sell error)
    ws.conditional_formatting.add(col_g, FormulaRule(
        formula=[f'AND(ISNUMBER(G{DS}),G{DS}<0)'],
        fill=_fill("FF6B6B")))

    # BUY → light green in column C
    ws.conditional_formatting.add(col_c, FormulaRule(
        formula=[f'$C{DS}="BUY"'], fill=F_LGRN))

    # SELL → light red in column C
    ws.conditional_formatting.add(col_c, FormulaRule(
        formula=[f'$C{DS}="SELL"'], fill=F_LRED))

    # P&L positive → green font in column L
    ws.conditional_formatting.add(col_l, FormulaRule(
        formula=[f'AND(ISNUMBER(L{DS}),L{DS}>0)'],
        font=Font(color="375623", bold=True)))

    # P&L negative → red font in column L
    ws.conditional_formatting.add(col_l, FormulaRule(
        formula=[f'AND(ISNUMBER(L{DS}),L{DS}<0)'],
        font=Font(color="9C0006", bold=True)))


# ─── Entry point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "grid_trading_calculator.xlsx")
    build(output_path)
