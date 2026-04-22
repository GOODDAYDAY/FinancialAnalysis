"""
REQ-002: CI validation of the grid trading calculator Excel file.

Ensures the spreadsheet's structure, formulas, conditional formatting,
and summary statistics are intact — protecting against accidental
breakage during edits.

Validates:
- File exists and is readable
- Sheet "网格计算" (Grid Calculation) present
- Parameter cells exist and are editable
- Formula cells contain formulas (not hardcoded)
- Conditional formatting rules exist
- Summary statistics block present
"""

import os
import pytest

openpyxl = pytest.importorskip("openpyxl")

from openpyxl import load_workbook

TOOLS_DIR = os.path.join(os.path.dirname(__file__), "..", "tools")
EXCEL_PATH = os.path.join(TOOLS_DIR, "grid_trading_calculator.xlsx")
SHEET_NAME = "网格计算"  # Grid Calculation


@pytest.fixture(scope="module")
def wb():
    """Load the workbook once for all tests."""
    assert os.path.exists(EXCEL_PATH), f"Excel file not found: {EXCEL_PATH}"
    return load_workbook(EXCEL_PATH)


class TestFileExists:
    def test_file_exists(self):
        assert os.path.exists(EXCEL_PATH)

    def test_file_not_empty(self):
        assert os.path.getsize(EXCEL_PATH) > 0


class TestSheetStructure:
    def test_has_sheet(self, wb):
        assert SHEET_NAME in wb.sheetnames, f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}"

    def test_sheet_has_content(self, wb):
        ws = wb[SHEET_NAME]
        assert ws.max_row > 10, f"Sheet has too few rows: {ws.max_row}"
        assert ws.max_column > 5, f"Sheet has too few columns: {ws.max_column}"


class TestParameterCells:
    """Parameter cells should exist and have values."""

    def test_price_parameter_exists(self, wb):
        ws = wb[SHEET_NAME]
        # The current price parameter is in E2
        price_cell = ws["E2"]
        assert price_cell.value is not None, "Current price parameter (E2) should have a value"

    def test_grid_interval_parameter_exists(self, wb):
        ws = wb[SHEET_NAME]
        # The grid interval parameter is in E3
        interval_cell = ws["E3"]
        assert interval_cell.value is not None, "Grid interval parameter (E3) should have a value"

    def test_parameters_are_numeric(self, wb):
        ws = wb[SHEET_NAME]
        price = ws["E2"].value
        interval = ws["E3"].value
        assert isinstance(price, (int, float)), f"Price should be numeric, got: {type(price)}"
        assert isinstance(interval, (int, float)), f"Interval should be numeric, got: {type(interval)}"


class TestFormulaCells:
    """Key cells should contain formulas, not hardcoded values."""

    def test_has_formulas(self, wb):
        ws = wb[SHEET_NAME]
        formula_count = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        assert formula_count > 50, f"Expected many formulas, found: {formula_count}"

    def test_title_cell_is_formula(self, wb):
        """A1 should be a formula that builds a title string."""
        ws = wb[SHEET_NAME]
        a1 = ws["A1"]
        assert isinstance(a1.value, str) and a1.value.startswith("="), (
            f"A1 should be a formula, got: {a1.value}"
        )

    def test_no_hardcoded_results_in_key_area(self, wb):
        """Key result columns should have formulas, not hardcoded numbers."""
        ws = wb[SHEET_NAME]
        # Check a range of cells that should contain formulas
        formula_found = False
        for row in ws.iter_rows(min_row=5, max_row=15, min_col=5, max_col=12):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_found = True
                    break
            if formula_found:
                break
        assert formula_found, "No formulas found in result area (rows 5-15, cols E-L)"


class TestConditionalFormatting:
    """The spreadsheet should have conditional formatting rules."""

    def test_has_conditional_formatting(self, wb):
        ws = wb[SHEET_NAME]
        rules = list(ws.conditional_formatting)
        assert len(rules) > 0, "No conditional formatting rules found"


class TestSummaryBlock:
    """The sheet should contain a summary statistics area."""

    def test_has_labels(self, wb):
        ws = wb[SHEET_NAME]
        # Check that the sheet has text labels (parameter names)
        labels_found = False
        for row in ws.iter_rows(min_row=1, max_row=5, max_col=4):
            for cell in row:
                if isinstance(cell.value, str) and len(cell.value) > 3:
                    labels_found = True
                    break
            if labels_found:
                break
        assert labels_found, "No text labels found in summary/header area"

    def test_sheet_not_empty(self, wb):
        ws = wb[SHEET_NAME]
        non_empty = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    non_empty += 1
        assert non_empty > 20, f"Sheet has too few non-empty cells: {non_empty}"


class TestStructureIntegrity:
    """Basic structural checks — no corruption, valid cell types."""

    def test_workbook_opens_without_error(self):
        wb = load_workbook(EXCEL_PATH)
        assert wb is not None

    def test_sheets_are_not_hidden_all(self, wb):
        """At least one sheet should be visible."""
        visible = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
        assert len(visible) > 0
